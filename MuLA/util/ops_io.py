import torch_geometric.transforms as T
from torch_geometric.datasets import Planetoid
import torch
import numpy as np
import scipy.sparse as sp
from util import ops_cvae_pt
from datasets.amazon import Amazon
from datasets.twitch import Twitch
from datasets.co_author import Coauthor
from datasets.lastfm_asia import LastFMAsia
from datasets.citation_full import CitationFull
from datasets.local_mat_data import LocalMatData
from datasets.attri_graph import AttributedGraphDataset
from datasets.wikipedia_network import WikipediaNetwork  
from datasets.wikics import WikiCS
import openpyxl

def get_pyg_dataset(direction, dataset_name, gnn_type):
    if dataset_name in ['CiteSeer', 'Cora', 'PubMed']:
        if gnn_type == 'gcn2':
            transform = T.Compose([T.NormalizeFeatures(), T.ToSparseTensor()])
        else:
            transform = T.NormalizeFeatures()
            # transform = T.Compose([T.NormalizeFeatures(), T.ToSparseTensor()])
        dataset = Planetoid(root=direction, name=dataset_name, transform=transform)
    elif dataset_name in ['Cora_Full', 'Cora_ML', 'DBLP']:
        if gnn_type == 'gcn2':
            transform = T.ToSparseTensor()
        else:
            transform = None
        dataset = CitationFull(root=direction, name=dataset_name, transform=transform)
    elif dataset_name in ['Computers', 'Photo']:
        if gnn_type == 'gcn2':
            transform = T.ToSparseTensor()
        else:
            transform = None
        dataset = Amazon(root=direction, name=dataset_name, transform=transform)
    elif dataset_name in ['CS', 'Physics']:
        if gnn_type == 'gcn2':
            transform = T.ToSparseTensor()
        else:
            transform = None
        dataset = Coauthor(root=direction, name=dataset_name, transform=transform)
    elif dataset_name in ['BlogCatalog', 'Flickr', 'Wiki']:
        if gnn_type == 'gcn2':
            transform = T.ToSparseTensor()
        else:
            transform = None
        dataset = AttributedGraphDataset(root=direction, name=dataset_name, transform=transform)
    elif dataset_name == 'ACM':
        if gnn_type == 'gcn2':
            transform = T.Compose([T.NormalizeFeatures(), T.ToSparseTensor()])
        else:
            transform = T.NormalizeFeatures()
        dataset = LocalMatData(root=direction, name='ACM', transform=transform)
    elif dataset_name == 'UAI':
        if gnn_type == 'gcn2':
            transform = T.ToSparseTensor()
        else:
            transform = None
        dataset = LocalMatData(root=direction, name='UAI', transform=transform)
    elif dataset_name in ['Twitch_DE', 'Twitch_EN', 'Twitch_ES', 'Twitch_FR', 'Twitch_PT', 'Twitch_RU']:
        if gnn_type == 'gcn2':
            transform = T.Compose([T.NormalizeFeatures(), T.ToSparseTensor()])
        else:
            transform = T.NormalizeFeatures()
        dataset = Twitch(root=direction, name=dataset_name.split('_')[1], transform=transform)
    elif dataset_name == 'LastFMAsia':
        if gnn_type == 'gcn2':
            transform = T.ToSparseTensor()
        else:
            transform = None
        dataset = LastFMAsia(root=direction, transform=transform)
    elif dataset_name == 'squirrel':
        if gnn_type == 'gcn2':
            transform = T.ToSparseTensor()
        else:
            transform = None
        dataset = WikipediaNetwork(root=direction, name='squirrel', transform=transform)
    elif dataset_name == 'WikiCS':
        if gnn_type == 'gcn2':
            transform = T.ToSparseTensor()
        else:
            transform = None
        dataset = WikiCS(root=direction, transform=transform)
    
    else:
        dataset = None

    return dataset


def normalize_features(features):
    """Row-normalize feature matrix and convert to tuple representation"""
    rowsum = np.array(features.sum(1))
    r_inv = np.power(rowsum, -1).flatten()
    r_inv[np.isinf(r_inv)] = 0.
    r_mat_inv = sp.diags(r_inv)
    features = r_mat_inv.dot(features)
    return features

def min_max_normalize(x):
    min_val = x.min()
    max_val = x.max()
    normalized_x = (x - min_val) / (max_val - min_val)
    return normalized_x

def get_augmented_features(load_saved, dataset_name, features, device, num_concat):
    if load_saved:
        cvae_model = torch.load("/home/gh/MuLA/CVAE_run_model/{}.pkl".format(dataset_name)).to(device)
    else:
        cvae_model = torch.load("/home/gh/MuLA/CVAE_run_model/{}.pkl".format(dataset_name)).to(device)

    x_list = []
    cvae_features = torch.tensor(features, dtype=torch.float32).to(device)
    for _ in range(num_concat):
        z = torch.randn([cvae_features.size(0), cvae_model.latent_size]).to(device)
        augmented_features = cvae_model.inference(z, cvae_features)
        augmented_features = ops_cvae_pt.feature_tensor_normalize(augmented_features).detach()
        x_list.append(augmented_features.to(device))

    return x_list



# def save_training_process(dataset_name, List_ID, List_loss, List_train, List_consis,List_mixup):
#     if len(List_ID) == 0:
#         pass
#     save_path = dataset_name + "_training_process.xlsx"
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = str(dataset_name)

#     # 设置列标题
#     headers = ["ID", "Loss", "loss_train", "loss_consis", "loss_mixup"]
#     for col_num, header in enumerate(headers, 1):
#         sheet.cell(row=1, column=col_num, value=header)

#     # 写入数据
#     for i in range(len(List_loss)):
#         sheet.cell(row=i + 2, column=1, value=str(List_ID[i])) # 从第二行开始写入数据
#         sheet.cell(row=i + 2, column=2, value=str(List_loss[i]))
#         sheet.cell(row=i + 2, column=3, value=str(List_train[i]))
#         sheet.cell(row=i + 2, column=4, value=str(List_consis[i]))
#         sheet.cell(row=i + 2, column=5, value=str(List_mixup[i]))
#         # sheet.cell(row=i + 2, column=13, value=str(List_un[i]))
#     workbook.save(save_path)
#     print("Finished the save of training process！")

def softmax(x):
    # 沿着第一个维度（axis=1）计算最大值，并广播减去它以提高数值稳定性
    maxes = np.max(x, axis=1, keepdims=True)
    e_x = np.exp(x - maxes)  # 更稳定的计算
    return e_x / np.sum(e_x, axis=1, keepdims=True)